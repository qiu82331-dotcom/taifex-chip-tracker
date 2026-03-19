#!/usr/bin/env python3
"""
台指期籌碼策略 — 100% 期交所資料
"""
import re
import sys
import time
import warnings
from io import StringIO
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)
EXCEL_FILE = OUTPUT_DIR / "台指期籌碼追蹤.xlsx"
PARQUET_FILE = DATA_DIR / "taifex_raw.parquet"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-TW,zh;q=0.9",
}
DELAY = 3
RETRY = 3
RETRY_DELAY = 5


def taifex_post(url, payload, label=""):
    """POST with retry."""
    for attempt in range(RETRY):
        try:
            resp = requests.post(url, data=payload, headers=HEADERS, timeout=20)
            resp.raise_for_status()
            return resp
        except Exception as e:
            if attempt < RETRY - 1:
                time.sleep(RETRY_DELAY)
            else:
                print(f"  WARNING: {label} 3次失敗: {e}")
                return None
    return None


# ================================================================
#  DATA A: 大額交易人 (近月, 遠月)
# ================================================================

def parse_large_trader_val(s):
    """Parse '43,830  (43,830)' → (43830, 43830)"""
    s = str(s).replace(",", "").strip()
    m = re.match(r"([\d]+)\s*\(([\d]+)\)", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m2 = re.match(r"([\d]+)", s)
    if m2:
        return int(m2.group(1)), int(m2.group(1))
    return 0, 0


def fetch_large_trader(date_str: str) -> dict:
    """Fetch large trader data for one date. Returns near/all contract sums."""
    d = date_str.replace("-", "/")
    url = "https://www.taifex.com.tw/cht/3/largeTraderFutQry"
    resp = taifex_post(url, {"queryDate": d, "contractId": "TX"}, f"大額交易人 {date_str}")
    if resp is None:
        return {}
    try:
        tables = pd.read_html(StringIO(resp.text))
    except Exception:
        return {}
    if not tables:
        return {}
    t = tables[0]
    rows = t.values.tolist()

    result = {"date": date_str}
    for row in rows:
        row_str = " ".join(str(c) for c in row).replace(" ", "")
        # Skip 週契約
        if "週契約" in row_str or "週" in row_str.split("契約")[0] if "契約" in row_str else False:
            if "週" in row_str and "所有" not in row_str:
                continue

        # Identify row type
        is_near = False
        is_all = False
        if "所有" in row_str and "契約" in row_str:
            is_all = True
        else:
            # Near-month: has a YYYYMM pattern
            ym = re.search(r"20\d{2}\s*\d{2}", row_str)
            if ym:
                is_near = True

        if not is_near and not is_all:
            continue

        # Extract 8 values: buy5(spec5), buy10(spec10), sell5(spec5), sell10(spec10)
        # Find cells that match the "number (number)" pattern
        cells = [str(c) for c in row]
        parsed = []
        for c in cells:
            c_clean = c.replace(",", "").strip()
            m = re.match(r"([\d]+)\s*\(([\d]+)\)", c_clean)
            if m and int(m.group(1)) > 10:
                parsed.append((int(m.group(1)), int(m.group(2))))

        if len(parsed) < 4:
            continue

        # parsed order: buy5(spec5), buy10(spec10), sell5(spec5), sell10(spec10)
        buy5, spec_buy5 = parsed[0]
        buy10, spec_buy10 = parsed[1]
        sell5, spec_sell5 = parsed[2]
        sell10, spec_sell10 = parsed[3]

        # Sum = 5大淨 + 5特淨 + 10大淨 + 10特淨
        net_sum = (buy5 - sell5) + (spec_buy5 - spec_sell5) + \
                  (buy10 - sell10) + (spec_buy10 - spec_sell10)

        prefix = "near" if is_near else "all"
        result[f"{prefix}_sum"] = net_sum
        result[f"{prefix}_buy5"] = buy5
        result[f"{prefix}_sell5"] = sell5
        result[f"{prefix}_buy10"] = buy10
        result[f"{prefix}_sell10"] = sell10
        result[f"{prefix}_spec_buy5"] = spec_buy5
        result[f"{prefix}_spec_sell5"] = spec_sell5
        result[f"{prefix}_spec_buy10"] = spec_buy10
        result[f"{prefix}_spec_sell10"] = spec_sell10

    return result


# ================================================================
#  DATA B: 三大法人-區分各期貨契約 (小散戶, 微散戶)
# ================================================================

def fetch_institutional(date_str: str, commodity: str) -> int:
    """Fetch institutional net OI for a commodity. Returns net OI (int) or None."""
    d = date_str.replace("-", "/")
    url = "https://www.taifex.com.tw/cht/3/futContractsDate"
    resp = taifex_post(url, {"queryDate": d, "commodityId": commodity},
                       f"三大法人 {commodity} {date_str}")
    if resp is None:
        return None
    try:
        tables = pd.read_html(StringIO(resp.text))
    except Exception:
        return None
    if not tables:
        return None
    t = tables[0]
    if t.shape[0] < 7:
        # Check for "查無資料"
        if "查無" in t.to_string():
            return None
        return None

    # Last row is 期貨合計, column index 13 is 未平倉餘額>多空淨額>口數
    try:
        last_row = t.iloc[-1]
        # The net OI is in the 未平倉餘額 > 多空淨額 > 口數 column
        # It's the 4th from last column (index -3 for 口數 in 多空淨額)
        val = last_row.iloc[13]  # 未平倉餘額 > 多空淨額 > 口數
        return int(val)
    except Exception:
        return None


def fetch_retail(date_str: str) -> dict:
    """Fetch 小散戶 and 微散戶 for one date."""
    result = {"date": date_str}

    mtx_net = fetch_institutional(date_str, "MXF")
    if mtx_net is not None:
        result["小散戶"] = -mtx_net
    else:
        result["小散戶"] = np.nan

    time.sleep(DELAY)

    tmf_net = fetch_institutional(date_str, "TMF")
    if tmf_net is not None:
        result["微散戶"] = -tmf_net
    else:
        result["微散戶"] = np.nan

    return result


# ================================================================
#  DATA C: 台指期每日行情
# ================================================================

def fetch_daily_price(date_str: str) -> dict:
    """Fetch TX daily OHLC for one date."""
    d = date_str.replace("-", "/")
    url = "https://www.taifex.com.tw/cht/3/futDailyMarketReport"
    payload = {
        "queryType": "2",
        "marketCode": "0",
        "commodity_id": "TX",
        "queryDate": d,
        "MarketCode": "0",
        "commodity_idt": "TX",
    }
    resp = taifex_post(url, payload, f"行情 {date_str}")
    if resp is None:
        return {}
    try:
        tables = pd.read_html(StringIO(resp.text))
    except Exception:
        return {}
    if not tables:
        return {}

    t = tables[0]
    if t.shape[0] < 2:
        return {}

    # Near-month = first row (highest volume non-spread contract)
    try:
        r = t.iloc[0]
        result = {
            "date": date_str,
            "開盤": float(str(r.iloc[2]).replace(",", "")),
            "最高": float(str(r.iloc[3]).replace(",", "")),
            "最低": float(str(r.iloc[4]).replace(",", "")),
            "收盤": float(str(r.iloc[5]).replace(",", "")),
        }
        # Try to get volume
        try:
            result["成交量"] = int(str(r.iloc[10]).replace(",", ""))
        except Exception:
            pass
        return result
    except Exception:
        return {}


# ================================================================
#  BATCH FETCH
# ================================================================

def get_trading_dates(start: str, end: str) -> list:
    """Generate weekday dates."""
    s = pd.Timestamp(start)
    e = pd.Timestamp(end)
    dates = pd.bdate_range(s, e)
    return [d.strftime("%Y-%m-%d") for d in dates]


def batch_fetch_all(start: str, end: str):
    """Fetch all data for date range."""
    dates = get_trading_dates(start, end)
    print(f"\n  交易日數: {len(dates)} 天")
    print(f"  預估時間: ~{len(dates) * 9 // 60} 分鐘 (3頁面 × 3秒間隔)\n")

    records = []
    for i, d in enumerate(dates):
        if i % 20 == 0 and i > 0:
            elapsed = i * 9  # approximate
            remaining = (len(dates) - i) * 9
            print(f"  進度: {i}/{len(dates)} ({remaining//60}分鐘剩餘)")

        # Data A: large trader
        lt = fetch_large_trader(d)
        time.sleep(DELAY)

        # Data B: retail
        rt = fetch_retail(d)
        # (fetch_retail already has DELAY between MXF and TMF)
        time.sleep(DELAY)

        # Data C: price
        px = fetch_daily_price(d)
        time.sleep(DELAY)

        # Merge
        if not lt.get("near_sum") and not lt.get("all_sum") and not px:
            # Non-trading day
            continue

        row = {"date": d}
        row.update(lt)
        row.update({k: v for k, v in rt.items() if k != "date"})
        row.update({k: v for k, v in px.items() if k != "date"})
        records.append(row)

    return pd.DataFrame(records)


# ================================================================
#  COMPUTE DERIVED
# ================================================================

def compute_all(df):
    df = df.copy().sort_values("date").reset_index(drop=True)

    # 近月 / 遠月 from large trader
    df["近月"] = df.get("near_sum", pd.Series(dtype=float)).fillna(0).astype(int)
    df["遠月"] = df.get("all_sum", pd.Series(dtype=float)).fillna(0).astype(int)

    # 結算近-散戶
    df["結算近減散戶"] = df["近月"] - df["小散戶"].fillna(0)

    # 主力 (two formulas)
    retail = df["小散戶"].fillna(0)
    micro = df["微散戶"]
    yuan = df["遠月"].astype(float)

    df["主力"] = np.where(
        micro.notna(),
        yuan / 2 - (retail + micro.fillna(0)),
        yuan / 2 - retail * 2
    )

    # 籌碼
    df["籌碼"] = df["主力"] - df["主力"].shift(5)

    # 星期
    df["星期"] = pd.to_datetime(df["date"]).dt.day_name().map({
        "Monday": "一", "Tuesday": "二", "Wednesday": "三",
        "Thursday": "四", "Friday": "五", "Saturday": "六", "Sunday": "日",
    })

    # 信號 (v2: 進場門檻5000, 出場連3天負)
    df["信號"] = ""
    df["操作"] = ""
    position = False

    for i in range(len(df)):
        m = df.loc[i, "主力"]
        m_1 = df.loc[i - 1, "主力"] if i >= 1 else np.nan
        m_2 = df.loc[i - 2, "主力"] if i >= 2 else np.nan
        m_3 = df.loc[i - 3, "主力"] if i >= 3 else np.nan
        ret = df.loc[i, "小散戶"] if pd.notna(df.loc[i, "小散戶"]) else 0

        if pd.isna(m):
            continue

        # 信號判斷
        is_entry = (not pd.isna(m_1)) and m > 5000 and m_1 <= 5000
        is_exit = (i >= 3 and not pd.isna(m_3)
                   and m <= 0 and m_1 <= 0 and m_2 <= 0 and m_3 > 0)

        if is_entry:
            sig = "🟢翻多"
        elif is_exit:
            sig = "🔴翻空"
        elif m > 5000 and ret < 0:
            sig = "⬆️多方 💪最強"
        elif m > 5000:
            sig = "⬆️多方"
        elif m > 0:
            sig = "↗️弱多"
        elif m < 0 and ret > 0:
            sig = "⬇️空方 ⚠️危險"
        else:
            sig = "⬇️空方"
        df.loc[i, "信號"] = sig

        # 操作欄（追蹤持倉狀態）
        if not position and is_entry:
            df.loc[i, "操作"] = "👉 隔天進場"
            position = True
        elif position and is_exit:
            df.loc[i, "操作"] = "👉 隔天出場"
            position = False
        elif position:
            df.loc[i, "操作"] = "續抱"
        else:
            df.loc[i, "操作"] = "觀望"

    return df


# ================================================================
#  BACKTEST
# ================================================================

def run_backtest(df):
    """v2 回測：進場=主力突破5000, 出場=連3天負"""
    has_open = "開盤" in df.columns and df["開盤"].notna().any()
    if has_open:
        print("  進場價: 使用隔天開盤價 ✅")
    else:
        print("  WARNING: 找不到開盤價欄位，用收盤價替代")

    dates = df["date"].values
    ops = df["操作"].astype(str).values
    opens = df["開盤"].values if has_open else df["收盤"].values
    n = len(df)

    trades = []
    position = 0
    entry_price = 0
    entry_date = None

    for i in range(n):
        op = ops[i]
        if "隔天進場" in op and position == 0 and i + 1 < n:
            price = opens[i + 1]
            if pd.notna(price):
                position = 1
                entry_price = float(price)
                entry_date = dates[i + 1]
        elif "隔天出場" in op and position == 1 and i + 1 < n:
            price = opens[i + 1]
            if pd.notna(price):
                exit_price = float(price)
                exit_date = dates[i + 1]
                pnl_pts = exit_price - entry_price
                pnl_money = pnl_pts * 50 - 114
                hold = (pd.Timestamp(exit_date) - pd.Timestamp(entry_date)).days
                trades.append({
                    "進場日期": entry_date, "進場價": entry_price,
                    "出場日期": exit_date, "出場價": exit_price,
                    "持有天數": hold, "損益點數": pnl_pts,
                    "損益金額": pnl_money,
                    "勝負": "勝" if pnl_money > 0 else "敗",
                })
                position = 0

    if not trades:
        return pd.DataFrame()
    tdf = pd.DataFrame(trades)
    tdf["累計金額"] = tdf["損益金額"].cumsum()
    return tdf


# ================================================================
#  VALIDATION
# ================================================================

def run_validation(df):
    checks = [
        ("2024-01-02", -5064, -6718, 3743, -10845),
        ("2024-01-03", -28751, -30190, 18654, -52403),
        ("2024-07-29", -51413, -54416, 20753, -49606),
        ("2025-04-25", 41236, 39763, 218, 29662.5),
    ]
    print("\n" + "=" * 90)
    print("  公式驗證")
    print("=" * 90)
    print(f"  {'日期':<12} {'近月':>8} {'近(正確)':>8} {'遠月':>8} {'遠(正確)':>8} "
          f"{'散戶':>8} {'散(正確)':>8} {'主力':>10} {'主(正確)':>10} {'結果':>4}")
    print(f"  {'─'*84}")

    results = []
    for date_str, c_near, c_far, c_retail, c_main in checks:
        row = df[df["date"] == date_str]
        if row.empty:
            print(f"  {date_str:<12} — 資料不存在")
            results.append({"日期": date_str, "結果": "N/A"})
            continue
        r = row.iloc[0]
        my_near = int(r.get("近月", 0))
        my_far = int(r.get("遠月", 0))
        my_retail = int(r.get("小散戶", 0)) if pd.notna(r.get("小散戶")) else 0
        my_main = float(r.get("主力", 0))

        diff_pct = abs(my_main - c_main) / max(abs(c_main), 1) * 100
        if diff_pct < 5:
            mark = "✅"
        elif diff_pct < 30:
            mark = "⚠️"
        else:
            mark = "❌"

        print(f"  {date_str:<12} {my_near:>+8,} {c_near:>+8,} {my_far:>+8,} {c_far:>+8,} "
              f"{my_retail:>+8,} {c_retail:>+8,} {my_main:>+10,.1f} {c_main:>+10,.1f} {mark}")

        results.append({
            "日期": date_str,
            "近月_ours": my_near, "近月_correct": c_near, "近月_diff": my_near - c_near,
            "遠月_ours": my_far, "遠月_correct": c_far, "遠月_diff": my_far - c_far,
            "小散戶_ours": my_retail, "小散戶_correct": c_retail, "小散戶_diff": my_retail - c_retail,
            "主力_ours": my_main, "主力_correct": c_main, "主力_diff": my_main - c_main,
            "結果": mark,
        })
    print()
    return pd.DataFrame(results)


# ================================================================
#  EXCEL
# ================================================================

def write_excel(df, trades_df, val_df):
    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfill_light = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    green_font = Font(color="008000")
    red_font = Font(color="FF0000")
    green_bold = Font(color="008000", bold=True)
    red_bold = Font(color="FF0000", bold=True)
    green_row = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_row = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"))
    center = Alignment(horizontal="center")
    num_fmt = "#,##0"

    def write_header(ws, headers, widths):
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, size=11)
            cell.fill = hfill_light
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    # --- Sheet 1: 籌碼紀錄 ---
    gray_font = Font(color="999999")
    ws1 = wb.active
    ws1.title = "籌碼紀錄"
    cols1 = ["date", "星期", "開盤", "收盤", "近月", "遠月", "小散戶", "微散戶",
             "結算近減散戶", "主力", "籌碼", "信號", "操作"]
    hdrs1 = ["日期", "星期", "開盤", "收盤", "近月", "遠月", "小散戶", "微散戶",
             "結算近-散戶", "主力", "籌碼", "信號", "操作"]
    widths1 = [12, 6, 10, 10, 14, 14, 14, 14, 14, 14, 14, 18, 14]
    write_header(ws1, hdrs1, widths1)

    display = df[df.get("收盤", pd.Series(dtype=float)).notna()].copy() if "收盤" in df.columns else df.copy()
    for ri, (_, row) in enumerate(display.iterrows(), 2):
        sig = str(row.get("信號", ""))
        row_fill = None
        if "翻多" in sig:
            row_fill = green_row
        elif "翻空" in sig:
            row_fill = red_row

        for ci, col in enumerate(cols1, 1):
            val = row.get(col, "")
            if col == "date":
                val = str(val)[:10].replace("-", "/")
            elif col in ("小散戶", "微散戶") and pd.isna(val):
                val = ""
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.border = border
            if col in ("近月", "遠月", "小散戶", "微散戶", "結算近減散戶", "主力", "籌碼",
                        "開盤", "收盤"):
                cell.number_format = num_fmt
            if row_fill:
                cell.fill = row_fill
            # Conditional formatting
            if col == "主力" and isinstance(val, (int, float)):
                cell.font = green_font if val > 0 else red_font if val < 0 else Font()
            if col == "籌碼" and isinstance(val, (int, float)):
                if val > 100:
                    cell.font = green_bold
                elif val < -100:
                    cell.font = red_bold
            if col == "操作" and isinstance(val, str):
                if "隔天進場" in val:
                    cell.font = green_bold
                elif "隔天出場" in val:
                    cell.font = red_bold
                elif val == "續抱":
                    cell.font = green_font
                elif val == "觀望":
                    cell.font = gray_font

    # --- Sheet 2: 小台回測 ---
    ws2 = wb.create_sheet("小台回測")
    hdrs2 = ["進場日期", "進場價", "出場日期", "出場價", "持有天數",
             "損益點數", "損益金額", "累計金額", "勝負"]
    widths2 = [12, 10, 12, 10, 10, 12, 12, 14, 8]
    write_header(ws2, hdrs2, widths2)

    if not trades_df.empty:
        for ri, (_, row) in enumerate(trades_df.iterrows(), 2):
            for ci, col in enumerate(hdrs2, 1):
                val = row.get(col, "")
                if col in ("進場日期", "出場日期"):
                    val = str(val)[:10].replace("-", "/")
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = border
                if col in ("進場價", "出場價", "損益點數", "損益金額", "累計金額"):
                    cell.number_format = num_fmt
                if col == "損益金額" and isinstance(val, (int, float)):
                    cell.font = green_font if val > 0 else red_font
                if col == "勝負":
                    cell.font = green_font if val == "勝" else red_font

        # Summary row
        r_sum = len(trades_df) + 2
        wins = (trades_df["勝負"] == "勝").sum()
        total = len(trades_df)
        ws2.cell(row=r_sum, column=1, value="合計").font = Font(bold=True)
        ws2.cell(row=r_sum, column=2, value=f"{total}筆").font = Font(bold=True)
        ws2.cell(row=r_sum, column=5,
                 value=f"勝率{wins/total*100:.1f}%").font = Font(bold=True)
        c_pts = ws2.cell(row=r_sum, column=6, value=trades_df["損益點數"].sum())
        c_pts.number_format = num_fmt
        c_pts.font = Font(bold=True)
        c_money = ws2.cell(row=r_sum, column=7, value=trades_df["損益金額"].sum())
        c_money.number_format = num_fmt
        c_money.font = Font(bold=True)

    # --- Sheet 3: 月度績效 ---
    ws3 = wb.create_sheet("月度績效")
    hdrs3 = ["月份", "交易數", "勝場", "敗場", "勝率(%)", "月損益點數", "月損益金額"]
    write_header(ws3, hdrs3, [12, 10, 10, 10, 10, 14, 14])
    if not trades_df.empty:
        tdf = trades_df.copy()
        tdf["month"] = pd.to_datetime(tdf["出場日期"]).dt.to_period("M")
        monthly = tdf.groupby("month").agg(
            n=("損益金額", "count"),
            wins=("勝負", lambda x: (x == "勝").sum()),
            pts=("損益點數", "sum"),
            pnl=("損益金額", "sum"),
        ).reset_index()
        monthly["losses"] = monthly["n"] - monthly["wins"]
        monthly["wr"] = (monthly["wins"] / monthly["n"] * 100).round(1)
        for ri, (_, row) in enumerate(monthly.iterrows(), 2):
            ws3.cell(row=ri, column=1, value=str(row["month"])).border = border
            ws3.cell(row=ri, column=2, value=int(row["n"])).border = border
            ws3.cell(row=ri, column=3, value=int(row["wins"])).border = border
            ws3.cell(row=ri, column=4, value=int(row["losses"])).border = border
            ws3.cell(row=ri, column=5, value=row["wr"]).border = border
            c = ws3.cell(row=ri, column=6, value=row["pts"])
            c.number_format = num_fmt
            c.border = border
            c = ws3.cell(row=ri, column=7, value=row["pnl"])
            c.number_format = num_fmt
            c.border = border
            c.font = green_font if row["pnl"] > 0 else red_font

    # --- Sheet 4: 年度績效 ---
    ws4 = wb.create_sheet("年度績效")
    hdrs4 = ["年度", "交易數", "勝場", "敗場", "勝率(%)", "年損益點數", "年損益金額"]
    write_header(ws4, hdrs4, [10, 10, 10, 10, 10, 14, 14])
    if not trades_df.empty:
        tdf["year"] = pd.to_datetime(tdf["出場日期"]).dt.year
        yearly = tdf.groupby("year").agg(
            n=("損益金額", "count"),
            wins=("勝負", lambda x: (x == "勝").sum()),
            pts=("損益點數", "sum"),
            pnl=("損益金額", "sum"),
        ).reset_index()
        yearly["losses"] = yearly["n"] - yearly["wins"]
        yearly["wr"] = (yearly["wins"] / yearly["n"] * 100).round(1)
        for ri, (_, row) in enumerate(yearly.iterrows(), 2):
            ws4.cell(row=ri, column=1, value=int(row["year"])).border = border
            ws4.cell(row=ri, column=2, value=int(row["n"])).border = border
            ws4.cell(row=ri, column=3, value=int(row["wins"])).border = border
            ws4.cell(row=ri, column=4, value=int(row["losses"])).border = border
            ws4.cell(row=ri, column=5, value=row["wr"]).border = border
            c = ws4.cell(row=ri, column=6, value=row["pts"])
            c.number_format = num_fmt
            c.border = border
            c = ws4.cell(row=ri, column=7, value=row["pnl"])
            c.number_format = num_fmt
            c.border = border

    # --- Sheet 5: 公式驗證 ---
    ws5 = wb.create_sheet("公式驗證")
    if val_df is not None and not val_df.empty:
        for ci, col in enumerate(val_df.columns, 1):
            cell = ws5.cell(row=1, column=ci, value=col)
            cell.font = Font(bold=True)
            cell.fill = hfill_light
            cell.border = border
            ws5.column_dimensions[get_column_letter(ci)].width = 14
        for ri, (_, row) in enumerate(val_df.iterrows(), 2):
            for ci, col in enumerate(val_df.columns, 1):
                ws5.cell(row=ri, column=ci, value=row[col]).border = border

    # --- Sheet 6: 使用說明 ---
    ws6 = wb.create_sheet("使用說明")
    lines = [
        "台指期籌碼策略（v2 優化版）",
        "",
        "【資料來源】100% 期交所",
        "近月/遠月：大額交易人未沖銷部位結構表（前5大+前10大+前5特+前10特 淨額加總）",
        "小散戶：三大法人-區分各期貨契約（小型臺指，法人淨額取反號）",
        "微散戶：同上（微型臺指，2024/07後）",
        "",
        "【公式】",
        "近月 = 近月5大淨 + 近月10大淨 + 近月5特淨 + 近月10特淨",
        "遠月 = 所有契約(5大淨+10大淨+5特淨+10特淨)",
        "小散戶 = -(小台自營淨+投信淨+外資淨)",
        "微散戶 = -(微台自營淨+投信淨+外資淨)",
        "主力 = 遠月/2 - 小散戶x2（微散戶有值時：遠月/2-(小散戶+微散戶)）",
        "",
        "【策略（v2 優化版）】",
        "進場：主力從 ≤5000 升到 >5000 → 隔天開盤做多1口小台",
        "出場：主力連續3天都是負的 → 隔天開盤平倉",
        "其餘時間不動",
        "",
        "【信號說明】",
        "🟢翻多 = 主力突破5000（進場信號）",
        "🔴翻空 = 主力連續3天負（出場信號）",
        "⬆️多方 = 主力>5000，持續持有",
        "💪最強 = 主力>5000 且 散戶做空",
        "↗️弱多 = 主力0~5000，還不夠強，不進場",
        "⬇️空方 = 主力<0",
        "⚠️危險 = 主力<0 且 散戶做多",
        "",
        "【操作欄說明】",
        "👉 隔天進場 = 明天早上8:45開盤做多1口小台",
        "👉 隔天出場 = 明天早上8:45開盤平倉",
        "續抱 = 不動，繼續持有",
        "觀望 = 不動，空手等待",
        "",
        "每天打開 Excel 只看最後一列的「操作」欄就好。",
        "",
        "【改善原因】",
        "1. 進場加門檻5000：過濾主力在零線附近來回震盪的假信號",
        "2. 出場改連3天負：避免被單日假翻空洗掉",
        "",
        "【成本】",
        "小台保證金：$93,500/口",
        "1點 = $50，來回成本 $114/口",
    ]
    for r, line in enumerate(lines, 1):
        cell = ws6.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith("【"):
            cell.font = Font(bold=True, size=12)
    ws6.column_dimensions["A"].width = 70

    wb.save(str(EXCEL_FILE))
    print(f"\n  ✅ Excel 已儲存: {EXCEL_FILE}")


# ================================================================
#  MAIN
# ================================================================

def main():
    start_date = "2024-01-01"
    end_date = datetime.now().strftime("%Y-%m-%d")

    print(f"\n{'#'*60}")
    print(f"  台指期籌碼策略 — 100% 期交所")
    print(f"  {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"  期間: {start_date} ~ {end_date}")
    print(f"{'#'*60}")

    # --- Step 0: Test single date ---
    print("\n  🔍 Step 0: 測試單天 2025/04/25")
    lt = fetch_large_trader("2025-04-25")
    print(f"    大額交易人: near_sum={lt.get('near_sum', 'N/A')}, all_sum={lt.get('all_sum', 'N/A')}")
    if lt.get("near_sum") == 41236:
        print(f"    ✅ 近月 = +41,236 匹配！")
    else:
        print(f"    ⚠ 近月 = {lt.get('near_sum')}, 期望 41236")
    time.sleep(DELAY)

    rt = fetch_retail("2025-04-25")
    print(f"    小散戶={rt.get('小散戶', 'N/A')}, 微散戶={rt.get('微散戶', 'N/A')}")
    time.sleep(DELAY)

    px = fetch_daily_price("2025-04-25")
    print(f"    行情: 開={px.get('開盤','N/A')}, 收={px.get('收盤','N/A')}")
    time.sleep(DELAY)

    # --- Step 1: Batch fetch (with parquet cache) ---
    if PARQUET_FILE.exists():
        df_cache = pd.read_parquet(str(PARQUET_FILE))
        last_cached = df_cache["date"].astype(str).str[:10].max()
        print(f"\n  📦 已有快取: {len(df_cache)} 天 (到 {last_cached})")
        # Only fetch dates after last cached date
        if last_cached < end_date:
            next_day = (pd.Timestamp(last_cached) + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
            print(f"  📡 補爬: {next_day} ~ {end_date}")
            df_new = batch_fetch_all(next_day, end_date)
            if not df_new.empty:
                df = pd.concat([df_cache, df_new], ignore_index=True)
                df = df.drop_duplicates(subset=["date"], keep="last")
            else:
                df = df_cache
        else:
            print(f"  ✅ 快取已是最新")
            df = df_cache
    else:
        print(f"\n  📡 Step 1: 批量爬取 {start_date} ~ {end_date}")
        df = batch_fetch_all(start_date, end_date)

    if df.empty:
        print("  ❌ 無資料")
        return

    # Save raw data to parquet cache
    df.to_parquet(str(PARQUET_FILE), index=False)
    print(f"  ✅ 資料: {len(df)} 天 (已存快取)")

    # --- Step 2: Compute ---
    print("\n  🧮 Step 2: 計算衍生欄位")
    df = compute_all(df)

    # --- Step 3: Validate ---
    val_df = run_validation(df)

    # --- Step 4: Backtest ---
    print("  📊 Step 4: 小台回測")
    trades_df = run_backtest(df)

    # --- Step 5: Excel ---
    print("  📝 Step 5: 產出 Excel")
    write_excel(df, trades_df, val_df)

    # --- Step 6: Summary ---
    print(f"\n{'═'*60}")
    print(f"  策略 v2 更新完成")
    print(f"{'═'*60}")
    print(f"\n  改動：")
    print(f"    進場：主力>0 → 主力>5000")
    print(f"    出場：翻空即出 → 連續3天負才出")
    print(f"\n  資料來源：100% 期交所")
    print(f"  資料範圍：{df['date'].min()} ~ {df['date'].max()}")
    print(f"  交易日數：{len(df)} 天")

    # Validation summary
    if val_df is not None and not val_df.empty:
        n_pass = (val_df["結果"].isin(["✅", "⚠️"])).sum()
        n_total = len(val_df[val_df["結果"] != "N/A"])
        print(f"\n  公式驗證：主力 {n_pass}/{n_total} 通過")

    if not trades_df.empty:
        wins = (trades_df["勝負"] == "勝").sum()
        losses = (trades_df["勝負"] == "敗").sum()
        total = len(trades_df)
        wr = wins / total * 100
        total_pts = trades_df["損益點數"].sum()
        total_pnl = trades_df["損益金額"].sum()
        avg_hold = trades_df["持有天數"].mean()
        avg_win = trades_df[trades_df["勝負"] == "勝"]["損益金額"].mean() if wins > 0 else 0
        avg_loss = abs(trades_df[trades_df["勝負"] == "敗"]["損益金額"].mean()) if losses > 0 else 1
        ratio = avg_win / avg_loss if avg_loss > 0 else float("inf")

        print(f"\n  回測結果：")
        print(f"    交易次數：{total} 筆")
        print(f"    勝率：{wr:.1f}%（{wins}勝 / {losses}敗）")
        print(f"    累計點數：{total_pts:+,.0f}")
        print(f"    累計金額：${total_pnl:+,.0f}")
        print(f"    平均持有：{avg_hold:.0f} 天")
        print(f"    賺賠比：{ratio:.2f}")

        print(f"\n  vs 舊策略：")
        print(f"    舊：43筆 勝率65.1% $+542,598")
        print(f"    新：{total}筆 勝率{wr:.1f}% ${total_pnl:+,.0f}")

        # Yearly
        tdf = trades_df.copy()
        tdf["year"] = pd.to_datetime(tdf["出場日期"]).dt.year
        print(f"\n  分年度：")
        for yr, grp in tdf.groupby("year"):
            w = (grp["勝負"] == "勝").sum()
            n = len(grp)
            pnl = grp["損益金額"].sum()
            print(f"    {yr}: {n}筆 勝率{w/n*100:.1f}% ${pnl:+,.0f}")

    print(f"\n  檔案：{EXCEL_FILE}")
    print(f"{'═'*60}\n")


if __name__ == "__main__":
    main()

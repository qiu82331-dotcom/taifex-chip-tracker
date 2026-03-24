#!/usr/bin/env python3
"""
每日盤後自動更新台指期籌碼 v2 — 100% 期交所
排程：每天 18:00 自動執行（macOS launchd）
手動：python scripts/update_today.py
"""
import sys
import time
import math
import warnings
from pathlib import Path
from datetime import date, datetime, timezone, timedelta

# 強制使用台灣時區，避免 GitHub Actions (UTC) 拿錯日期
TW_TZ = timezone(timedelta(hours=8))

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = ROOT / "output" / "台指期籌碼追蹤.xlsx"
CSV_FILE = ROOT / "data" / "chip_history.csv"
PARQUET_FILE = ROOT / "data" / "taifex_raw.parquet"
LOG_DIR = ROOT / "logs"
LOG_DIR.mkdir(exist_ok=True)
(ROOT / "data").mkdir(exist_ok=True)
(ROOT / "output").mkdir(exist_ok=True)

# Import fetch functions from taifex_strategy
sys.path.insert(0, str(ROOT / "scripts"))
from taifex_strategy import fetch_large_trader, fetch_retail, fetch_daily_price, DELAY


def main():
    today = datetime.now(TW_TZ).date()
    today_str = today.strftime("%Y-%m-%d")
    print(f"\n{'='*50}")
    print(f"  台指期籌碼更新 v2 — {today_str}")
    print(f"{'='*50}")

    # ── Step 1: 讀取現有 Excel ──
    if not EXCEL_FILE.exists():
        print(f"  找不到 Excel: {EXCEL_FILE}")
        print(f"  請先跑: python scripts/taifex_strategy.py")
        sys.exit(1)

    wb = load_workbook(str(EXCEL_FILE))
    ws = wb["籌碼紀錄"]

    # 找最後一列有日期的 row
    last_row = ws.max_row
    while last_row > 1:
        cell_val = ws.cell(row=last_row, column=1).value
        if cell_val is not None and str(cell_val).strip() != "" and "合計" not in str(cell_val):
            break
        last_row -= 1

    last_date = ws.cell(row=last_row, column=1).value

    # 比對是否已更新
    if isinstance(last_date, datetime):
        last_date_cmp = last_date.date()
    elif isinstance(last_date, str):
        last_date_cmp = datetime.strptime(last_date.replace("-", "/"), "%Y/%m/%d").date()
    else:
        last_date_cmp = None

    if last_date_cmp == today:
        print(f"  今天 {today} 已更新過，跳過")
        sys.exit(0)

    # 讀取最近幾天主力值（column 10）用於信號判斷
    def read_main(row_idx):
        v = ws.cell(row=row_idx, column=10).value
        return float(v) if v is not None else 0.0

    m_1ago = read_main(last_row)          # 昨天
    m_2ago = read_main(max(last_row - 1, 2))  # 前天
    m_3ago = read_main(max(last_row - 2, 2))  # 大前天

    # 讀取5天前主力（column 10）→ 給籌碼計算用
    main_5d_ago = read_main(max(last_row - 4, 2))

    # 讀取昨天操作欄（column 15）判斷是否持倉
    last_op = ws.cell(row=last_row, column=15).value or ""
    is_holding = last_op in ("續抱", "👉 隔天進場")

    print(f"  最後日期：{last_date}")
    print(f"  昨天主力：{m_1ago:+,.1f}")
    print(f"  持倉狀態：{'持有中' if is_holding else '空手'}")

    # ── Step 2: 判斷交易日 ──
    weekday = today.weekday()
    if weekday >= 5:
        day_name = "六" if weekday == 5 else "日"
        print(f"  今天是週{day_name}，非交易日，跳過")
        sys.exit(0)

    # ── Step 3: 爬大額交易人（近月、遠月）──
    print(f"\n  爬取大額交易人...")
    lt = fetch_large_trader(today_str)
    if not lt or ("near_sum" not in lt and "all_sum" not in lt):
        print(f"  大額交易人爬取失敗或無資料（可能為國定假日或期交所尚未更新）")
        sys.exit(1)

    near_sum = lt.get("near_sum", 0)
    all_sum = lt.get("all_sum", 0)
    print(f"    近月(near_sum)={near_sum:+,}, 全部(all_sum)={all_sum:+,}")
    time.sleep(DELAY)

    # ── Step 4: 爬三大法人（小散戶、微散戶）──
    print(f"  爬取三大法人...")
    rt = fetch_retail(today_str)
    小散戶 = rt.get("小散戶", np.nan)
    微散戶 = rt.get("微散戶", np.nan)
    print(f"    小散戶={小散戶}, 微散戶={微散戶}")
    time.sleep(DELAY)

    # ── Step 5: 爬台指期行情 ──
    print(f"  爬取台指期行情...")
    px = fetch_daily_price(today_str)
    if not px or "開盤" not in px:
        print(f"  行情爬取失敗或無資料")
        sys.exit(1)

    開盤 = px["開盤"]
    收盤 = px["收盤"]
    print(f"    開盤={開盤:,.0f}, 收盤={收盤:,.0f}")

    # ── Step 6: 計算衍生欄位 ──
    近月 = near_sum
    遠月 = all_sum
    結算近減散戶 = 近月 - (小散戶 if not math.isnan(小散戶) else 0)

    if not math.isnan(微散戶):
        主力 = 遠月 / 2 - (小散戶 + 微散戶)
    else:
        主力 = 遠月 / 2 - 小散戶 * 2

    籌碼 = 主力 - main_5d_ago

    # v2 信號判斷
    is_entry = 主力 > 5000 and m_1ago <= 5000
    is_exit = (主力 <= 0 and m_1ago <= 0 and m_2ago <= 0 and m_3ago > 0)
    ret = 小散戶 if not math.isnan(小散戶) else 0

    if is_entry:
        信號 = "🟢翻多"
    elif is_exit:
        信號 = "🔴翻空"
    elif 主力 > 5000 and ret < 0:
        信號 = "⬆️多方 💪最強"
    elif 主力 > 5000:
        信號 = "⬆️多方"
    elif 主力 > 0:
        信號 = "↗️弱多"
    elif 主力 < 0 and ret > 0:
        信號 = "⬇️空方 ⚠️危險"
    else:
        信號 = "⬇️空方"

    # v2 操作欄
    if not is_holding and is_entry:
        操作 = "👉 隔天進場"
    elif is_holding and is_exit:
        操作 = "👉 隔天出場"
    elif is_holding:
        操作 = "續抱"
    else:
        操作 = "觀望"

    weekday_map = {0: "一", 1: "二", 2: "三", 3: "四", 4: "五"}
    星期 = weekday_map[today.weekday()]

    # ── Step 6b: 計算未實現損益 & 最低保證金 ──
    未實現損益 = ""
    最低保證金 = ""
    if 操作 == "👉 隔天進場":
        未實現損益 = 0
        最低保證金 = 71750
    elif 操作 in ("續抱", "👉 隔天出場"):
        # 往上掃描找最近的「👉 隔天進場」，進場價 = 那列下一列的開盤
        entry_row = last_row
        while entry_row > 1:
            op_val = ws.cell(row=entry_row, column=15).value or ""
            if op_val == "👉 隔天進場":
                break
            entry_row -= 1
        entry_price = ws.cell(row=entry_row + 1, column=3).value or 0
        pnl = 收盤 - entry_price
        未實現損益 = pnl
        最低保證金 = 71750 + abs(pnl) * 50 if pnl < 0 else 71750

    # ── Step 7: 寫入 Excel ──
    new_row = last_row + 1
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_fmt = "#,##0"
    green_font = Font(color="008000")
    red_font = Font(color="FF0000")
    green_bold = Font(color="008000", bold=True)
    red_bold = Font(color="FF0000", bold=True)
    gray_font = Font(color="999999")

    ws.cell(row=new_row, column=1, value=today.strftime("%Y/%m/%d"))
    ws.cell(row=new_row, column=2, value=星期)
    ws.cell(row=new_row, column=3, value=開盤)
    ws.cell(row=new_row, column=4, value=收盤)
    ws.cell(row=new_row, column=5, value=近月)
    ws.cell(row=new_row, column=6, value=遠月)
    ws.cell(row=new_row, column=7, value=小散戶 if not math.isnan(小散戶) else "")
    ws.cell(row=new_row, column=8, value=微散戶 if not math.isnan(微散戶) else "")
    ws.cell(row=new_row, column=9, value=結算近減散戶)
    ws.cell(row=new_row, column=10, value=主力)
    ws.cell(row=new_row, column=11, value=籌碼)
    ws.cell(row=new_row, column=12, value=未實現損益)
    ws.cell(row=new_row, column=13, value=最低保證金)
    ws.cell(row=new_row, column=14, value=信號)
    ws.cell(row=new_row, column=15, value=操作)

    # 邊框 + 數字格式
    for col in range(1, 16):
        ws.cell(row=new_row, column=col).border = border
    for col in [3, 4, 5, 6, 7, 8, 9, 10, 11]:
        cell = ws.cell(row=new_row, column=col)
        if cell.value != "":
            cell.number_format = num_fmt
    if 未實現損益 != "":
        ws.cell(row=new_row, column=12).number_format = num_fmt
    if 最低保證金 != "":
        ws.cell(row=new_row, column=13).number_format = "$#,##0"

    # 主力顏色
    if 主力 > 0:
        ws.cell(row=new_row, column=10).font = green_font
    elif 主力 < 0:
        ws.cell(row=new_row, column=10).font = red_font

    # 籌碼顏色
    if 籌碼 > 100:
        ws.cell(row=new_row, column=11).font = green_bold
    elif 籌碼 < -100:
        ws.cell(row=new_row, column=11).font = red_bold

    # 未實現損益顏色
    if isinstance(未實現損益, (int, float)):
        if 未實現損益 < -300:
            ws.cell(row=new_row, column=12).font = red_bold
        elif 未實現損益 < 0:
            ws.cell(row=new_row, column=12).font = red_font
        elif 未實現損益 > 0:
            ws.cell(row=new_row, column=12).font = green_font

    # 最低保證金顏色
    if isinstance(最低保證金, (int, float)):
        if 最低保證金 > 120000:
            ws.cell(row=new_row, column=13).font = red_font
        else:
            ws.cell(row=new_row, column=13).font = green_font

    # 操作欄顏色
    if "隔天進場" in 操作:
        ws.cell(row=new_row, column=15).font = green_bold
    elif "隔天出場" in 操作:
        ws.cell(row=new_row, column=15).font = red_bold
    elif 操作 == "續抱":
        ws.cell(row=new_row, column=15).font = green_font
    elif 操作 == "觀望":
        ws.cell(row=new_row, column=15).font = gray_font

    # 翻多翻空整列底色
    if "翻多" in 信號:
        fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for col in range(1, 16):
            ws.cell(row=new_row, column=col).fill = fill
    elif "翻空" in 信號:
        fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        for col in range(1, 16):
            ws.cell(row=new_row, column=col).fill = fill

    wb.save(str(EXCEL_FILE))

    # ── Step 7b: 更新 parquet 快取 ──
    new_record = {
        "date": today_str, "開盤": 開盤, "收盤": 收盤,
        "near_sum": 近月, "all_sum": 遠月,
        "小散戶": 小散戶 if not math.isnan(小散戶) else None,
        "微散戶": 微散戶 if not math.isnan(微散戶) else None,
    }
    if PARQUET_FILE.exists():
        df_cache = pd.read_parquet(str(PARQUET_FILE))
        df_cache = df_cache[df_cache["date"].astype(str).str[:10] != today_str]
        df_cache = pd.concat([df_cache, pd.DataFrame([new_record])], ignore_index=True)
        df_cache.to_parquet(str(PARQUET_FILE), index=False)

    # ── Step 7c: 產出 chip_history.csv（給 Google Sheet 讀）──
    df_excel = pd.read_excel(str(EXCEL_FILE), sheet_name="籌碼紀錄")
    df_excel["日期"] = df_excel["日期"].astype(str).str[:10]
    df_excel.to_csv(str(CSV_FILE), index=False, encoding="utf-8-sig")
    print(f"  CSV 已更新: {CSV_FILE}")

    # ── Step 8: 印出結果 ──
    print(f"""
═══════════════════════════════
  {today.strftime('%Y/%m/%d')} ({星期}) 更新完成
  收盤：{收盤:,.0f}
  近月：{近月:+,}
  遠月：{遠月:+,}
  小散戶：{小散戶:+,.0f}
  主力：{主力:+,.1f}
  籌碼：{籌碼:+,.1f}
  信號：{信號}
  操作：{操作}
  已更新到 {EXCEL_FILE}
═══════════════════════════════""")

    # 特別提醒
    if "隔天進場" in 操作:
        print("\n  🟢🟢🟢 明天早上8:45開盤做多1口小台！🟢🟢🟢")
    elif "隔天出場" in 操作:
        print("\n  🔴🔴🔴 明天早上8:45開盤平倉！🔴🔴🔴")


if __name__ == "__main__":
    main()

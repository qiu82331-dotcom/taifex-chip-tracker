#!/usr/bin/env python3
"""
小台指期貨籌碼策略 — 從零開始
用法:
  python scripts/chip_strategy.py              # 完整執行
  python scripts/chip_strategy.py --verify     # 只跑驗證步驟
  python scripts/chip_strategy.py --update     # 每日更新
"""

import sys
import time
import requests
import warnings
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# === 路徑 ===
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
PARQUET_FILE = DATA_DIR / "chip_data.parquet"
EXCEL_FILE = OUTPUT_DIR / "台指期籌碼追蹤.xlsx"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# === FinMind ===
FINMIND_API = "https://api.finmindtrade.com/api/v4/data"
FINMIND_TOKEN = ""  # 免費即可，有 token 可避免限速
REQUEST_DELAY = 0.5

# === 小台參數 ===
MTX_POINT_VALUE = 50       # 1點 = $50
MTX_MARGIN = 93_500        # 保證金
MTX_COST = 114             # 來回手續費+稅
MTX_LOTS = 1               # 固定1口


# =============================================
#  STEP 1: FinMind 資料抓取 (未平倉 OI)
# =============================================

def finmind_fetch(dataset: str, params: dict) -> pd.DataFrame:
    """通用 FinMind REST API 呼叫"""
    base = {
        "dataset": dataset,
        "token": FINMIND_TOKEN,
    }
    base.update(params)
    resp = requests.get(FINMIND_API, params=base, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") != 200 or "data" not in data:
        print(f"  ⚠ FinMind {dataset} 回傳異常: {data.get('msg', '')}")
        return pd.DataFrame()
    return pd.DataFrame(data["data"])


def fetch_futures_institutional(start: str, end: str) -> pd.DataFrame:
    """
    抓三大法人台指期未平倉 (OI)
    dataset: TaiwanFuturesInstitutionalInvestors
    關鍵欄位:
      - long_open_interest_balance_volume  (多方未平倉口數)
      - short_open_interest_balance_volume (空方未平倉口數)
    ⚠ 不要用 long_deal_volume / short_deal_volume — 那是買賣超(交易量)！
    """
    print(f"  抓取三大法人台指期 OI: {start} ~ {end}")
    df = finmind_fetch("TaiwanFuturesInstitutionalInvestors", {
        "data_id": "TX",
        "start_date": start,
        "end_date": end,
    })
    if df.empty:
        return df

    # === STEP 1 驗證：印出欄位名稱 ===
    print(f"\n  📋 原始欄位: {list(df.columns)}")
    print(f"  📋 institutional_investors 類別: {df['institutional_investors'].unique().tolist()}")
    print(f"  📋 前5筆:")
    print(df.head().to_string(index=False))

    # 確認是未平倉欄位
    assert "long_open_interest_balance_volume" in df.columns, \
        "❌ 找不到 long_open_interest_balance_volume！可能抓到買賣超了"
    assert "short_open_interest_balance_volume" in df.columns, \
        "❌ 找不到 short_open_interest_balance_volume！"

    return df


def fetch_futures_daily(start: str, end: str) -> pd.DataFrame:
    """抓台指期每日行情 (收盤價、成交量)"""
    print(f"  抓取台指期行情: {start} ~ {end}")
    df = finmind_fetch("TaiwanFuturesDaily", {
        "data_id": "TX",
        "start_date": start,
        "end_date": end,
    })
    if df.empty:
        return df
    # 只取近月 (contract_date 最小的)
    df["contract_date"] = df["contract_date"].astype(str)
    df = df[df["contract_date"].str.len() == 6]  # YYYYMM format
    result = []
    for date_val, grp in df.groupby("date"):
        nearest = grp.sort_values("contract_date").iloc[0]
        result.append({
            "date": date_val,
            "futures_close": float(nearest["close"]),
            "futures_open": float(nearest["open"]),
            "futures_volume": int(nearest.get("volume", 0)),
        })
    return pd.DataFrame(result)


def fetch_mini_futures_institutional(start: str, end: str) -> pd.DataFrame:
    """抓小台(MTX)三大法人未平倉"""
    print(f"  抓取小台三大法人 OI: {start} ~ {end}")
    df = finmind_fetch("TaiwanFuturesInstitutionalInvestors", {
        "data_id": "MTX",
        "start_date": start,
        "end_date": end,
    })
    return df


def fetch_option_institutional(start: str, end: str) -> pd.DataFrame:
    """抓選擇權三大法人 (for P/C ratio)"""
    print(f"  抓取選擇權三大法人: {start} ~ {end}")
    df = finmind_fetch("TaiwanOptionInstitutionalInvestors", {
        "data_id": "TXO",
        "start_date": start,
        "end_date": end,
    })
    return df


def fetch_taiex(start: str, end: str) -> pd.DataFrame:
    """抓加權指數"""
    print(f"  抓取加權指數: {start} ~ {end}")
    df = finmind_fetch("TaiwanStockTotalReturn", {
        "data_id": "TAIEX",
        "start_date": start,
        "end_date": end,
    })
    if df.empty:
        # fallback: 用 TaiwanStockPrice for ^TWII
        df = finmind_fetch("TaiwanVariousIndicators5Seconds", {
            "start_date": start,
            "end_date": end,
        })
    return df


def fetch_foreign_spot(start: str, end: str) -> pd.DataFrame:
    """抓外資現貨買賣超"""
    print(f"  抓取外資現貨買賣超: {start} ~ {end}")
    df = finmind_fetch("TaiwanStockMarginPurchaseShortSale", {
        "start_date": start,
        "end_date": end,
    })
    return df


# =============================================
#  STEP 2: TAIFEX 大額交易人 (for 主力)
# =============================================

def fetch_large_traders_taifex(date_str: str) -> dict:
    """
    從期交所爬大額交易人未沖銷部位
    URL: https://www.taifex.com.tw/cht/3/largeTraderFutQry
    回傳: top5_buy, top5_sell, top10_buy, top10_sell (所有契約)
    """
    import re
    url = "https://www.taifex.com.tw/cht/3/largeTraderFutQry"
    d = date_str.replace("-", "/")
    payload = {
        "queryDate": d,
        "contractId": "TX",
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": url,
    }
    try:
        resp = requests.post(url, data=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        html = resp.text

        # 用 pandas 解析 HTML 表格
        tables = pd.read_html(html, encoding="utf-8")
        if not tables:
            return {}

        # 找到正確的表格 (含 "所有契約")
        target = None
        for t in tables:
            t_str = t.to_string()
            if "買方" in t_str and "賣方" in t_str:
                target = t
                break
        if target is None:
            return {}

        # 解析數值 — 格式: "50,518(50,518)"
        def parse_val(s):
            s = str(s).replace(",", "").strip()
            # 取括號外的數字 (所有契約)
            m = re.match(r"([\d]+)", s)
            return int(m.group(1)) if m else 0

        def parse_special(s):
            """取括號內的數字 (特定法人)"""
            s = str(s).replace(",", "").strip()
            m = re.search(r"\(([\d]+)\)", s)
            return int(m.group(1)) if m else 0

        # 所有契約那列 (通常最後幾行)
        rows = target.values.tolist()

        # 找 "所有契約" 的行
        all_contract_rows = []
        for i, row in enumerate(rows):
            row_str = " ".join(str(c) for c in row)
            if "所有" in row_str or "all" in row_str.lower():
                all_contract_rows.append(row)

        if not all_contract_rows:
            # fallback: 最後一行通常是所有契約
            all_contract_rows = [rows[-1]]

        # 解析: 5大買, 10大買, 5大賣, 10大賣
        result = {"date": date_str}
        for row in all_contract_rows:
            nums = []
            for cell in row:
                s = str(cell).replace(",", "")
                m = re.match(r"([\d]+)", s)
                if m and int(m.group(1)) > 100:
                    nums.append(int(m.group(1)))
                    # 也取特定法人
                    m2 = re.search(r"\(([\d]+)\)", s)
                    if m2:
                        nums.append(int(m2.group(1)))

            if len(nums) >= 8:
                result["all_top5_buy"] = nums[0]
                result["spec_top5_buy"] = nums[1]
                result["all_top10_buy"] = nums[2]
                result["spec_top10_buy"] = nums[3]
                result["all_top5_sell"] = nums[4]
                result["spec_top5_sell"] = nums[5]
                result["all_top10_sell"] = nums[6]
                result["spec_top10_sell"] = nums[7]
                if len(nums) >= 9:
                    result["total_oi"] = nums[8]
                break

        return result if len(result) > 2 else {}

    except Exception as e:
        print(f"  ⚠ TAIFEX 大額交易人爬取失敗 {date_str}: {e}")
        return {}


def batch_large_traders(dates: list) -> pd.DataFrame:
    """批次爬大額交易人"""
    print(f"  批次爬取大額交易人: {len(dates)} 天")
    records = []
    for i, d in enumerate(dates):
        if i > 0 and i % 10 == 0:
            print(f"    進度: {i}/{len(dates)}")
        lt = fetch_large_traders_taifex(d)
        if lt:
            records.append(lt)
        time.sleep(REQUEST_DELAY)
    return pd.DataFrame(records) if records else pd.DataFrame()


# =============================================
#  STEP 3: 資料整合 + 衍生欄位計算
# =============================================

def process_institutional(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    把三大法人原始資料轉成每日一行
    關鍵: 用 long_open_interest_balance_volume / short_open_interest_balance_volume (未平倉!)
    """
    if df_raw.empty:
        return pd.DataFrame()

    records = []
    for date_val, grp in df_raw.groupby("date"):
        row = {"date": date_val}
        for _, r in grp.iterrows():
            name = r["institutional_investors"]
            long_oi = int(r["long_open_interest_balance_volume"])
            short_oi = int(r["short_open_interest_balance_volume"])
            net_oi = long_oi - short_oi

            if "外資" in name or "Foreign" in name:
                row["foreign_long_oi"] = long_oi
                row["foreign_short_oi"] = short_oi
                row["foreign_net_oi"] = net_oi
            elif "投信" in name or "Trust" in name:
                row["trust_long_oi"] = long_oi
                row["trust_short_oi"] = short_oi
                row["trust_net_oi"] = net_oi
            elif "自營" in name or "Dealer" in name:
                row["dealer_long_oi"] = long_oi
                row["dealer_short_oi"] = short_oi
                row["dealer_net_oi"] = net_oi
        records.append(row)
    return pd.DataFrame(records)


def process_mini_institutional(df_raw: pd.DataFrame) -> pd.DataFrame:
    """小台三大法人未平倉"""
    if df_raw.empty:
        return pd.DataFrame()

    records = []
    for date_val, grp in df_raw.groupby("date"):
        row = {"date": date_val}
        total_long = 0
        total_short = 0
        for _, r in grp.iterrows():
            long_oi = int(r.get("long_open_interest_balance_volume", 0))
            short_oi = int(r.get("short_open_interest_balance_volume", 0))
            total_long += long_oi
            total_short += short_oi
        row["mtx_inst_long_oi"] = total_long
        row["mtx_inst_short_oi"] = total_short
        row["mtx_inst_net_oi"] = total_long - total_short
        records.append(row)
    return pd.DataFrame(records)


def compute_derived(df: pd.DataFrame) -> pd.DataFrame:
    """
    計算衍生欄位 (v2 修正版):
    - 遠月 = inst_total (三大法人全合約淨OI，替代真正的遠月合約OI)
    - 小散戶 = -(小台法人淨未平倉) — 不除以4！
    - 主力 = 遠月/2 - 小散戶×2 (新公式)
    - 籌碼 = 主力 - 主力.shift(5) (5日動量)
    - 信號: 翻多(主力由負翻正), 翻空(主力由正翻負)
    """
    df = df.copy()
    df = df.sort_values("date").reset_index(drop=True)

    # === 遠月 = 三大法人全合約淨OI (inst_total) ===
    # 注意: TAIFEX 無法取得近月/遠月分拆，用全合約合計替代
    foreign = df.get("foreign_net_oi", pd.Series(0, index=df.index)).fillna(0)
    trust = df.get("trust_net_oi", pd.Series(0, index=df.index)).fillna(0)
    dealer = df.get("dealer_net_oi", pd.Series(0, index=df.index)).fillna(0)
    df["遠月"] = foreign + trust + dealer  # inst_total

    # === 小散戶 = -(小台法人淨OI) — 不除以4 ===
    if "mtx_inst_net_oi" in df.columns:
        df["小散戶"] = -(df["mtx_inst_net_oi"].fillna(0)).astype(int)
    else:
        df["小散戶"] = 0

    # === 主力 = 遠月/2 - 小散戶×2 (新公式) ===
    df["主力"] = (df["遠月"] / 2 - df["小散戶"] * 2).round(0).astype(int)

    # === 近月 (外資淨OI，保留參考) ===
    df["近月"] = foreign.astype(int)

    # === 結算近-散戶 ===
    df["結算近-散戶"] = df["近月"] - df["小散戶"]

    # === 籌碼 = 主力 - 主力.shift(5) (5日動量) ===
    df["籌碼"] = df["主力"] - df["主力"].shift(5).fillna(0).astype(int)

    # === 背離 ===
    if "futures_close" in df.columns:
        price_chg = df["futures_close"].diff()
        df["背離"] = ""
        for i in range(1, len(df)):
            if df.loc[i, "主力"] > 0 and price_chg.iloc[i] < 0:
                df.loc[i, "背離"] = "多方背離"
            elif df.loc[i, "主力"] < 0 and price_chg.iloc[i] > 0:
                df.loc[i, "背離"] = "空方背離"
    else:
        df["背離"] = ""

    # === 信號 ===
    df["信號"] = ""
    for i in range(1, len(df)):
        prev_main = df.loc[i - 1, "主力"]
        curr_main = df.loc[i, "主力"]
        curr_retail = df.loc[i, "小散戶"]

        if prev_main <= 0 and curr_main > 0:
            sig = "🔼翻多"
        elif prev_main >= 0 and curr_main < 0:
            sig = "🔽翻空"
        elif curr_main > 0:
            sig = "多方"
        elif curr_main < 0:
            sig = "空方"
        else:
            sig = "—"

        # 加強信號
        if curr_main > 0 and curr_retail < 0:
            sig += " 💪主多散空"
        elif curr_main < 0 and curr_retail > 0:
            sig += " ⚠️主空散多"

        df.loc[i, "信號"] = sig

    return df


# =============================================
#  STEP 4: 小台回測
# =============================================

def run_backtest(df: pd.DataFrame) -> pd.DataFrame:
    """
    小台回測 (v2 修正: 用隔天開盤價進出場):
    - 翻多信號日 → 隔天開盤買進
    - 翻空信號日 → 隔天開盤賣出
    - 固定1口, 來回成本 $114
    """
    if "信號" not in df.columns or "futures_close" not in df.columns:
        return pd.DataFrame()

    # 檢查是否有開盤價欄位
    has_open = "futures_open" in df.columns and df["futures_open"].notna().any()
    if has_open:
        open_vals = df["futures_open"].values
        print("  回測進出場價: 使用隔天開盤價 ✅")
    else:
        open_vals = df["futures_close"].values
        print("  WARNING: 找不到開盤價欄位，用收盤價替代")

    dates = df["date"].values
    signals = df["信號"].astype(str).values
    n = len(df)

    trades = []
    position = None  # None or {"signal_idx", ...}

    for i in range(n):
        sig = signals[i]

        if "翻多" in sig and position is None:
            # 隔天開盤買進
            if i + 1 < n:
                entry_price = float(open_vals[i + 1])
                position = {
                    "signal_date": dates[i],
                    "entry_date": dates[i + 1],
                    "entry_price": entry_price,
                }

        elif "翻空" in sig and position is not None:
            # 隔天開盤賣出
            if i + 1 < n:
                exit_price = float(open_vals[i + 1])
                exit_date = dates[i + 1]
            else:
                # 最後一天用當天收盤
                exit_price = float(df["futures_close"].iloc[i])
                exit_date = dates[i]

            pnl_points = exit_price - position["entry_price"]
            pnl_money = pnl_points * MTX_POINT_VALUE * MTX_LOTS - MTX_COST
            trades.append({
                "進場日": position["entry_date"],
                "進場價": position["entry_price"],
                "出場日": exit_date,
                "出場價": exit_price,
                "點數損益": pnl_points,
                "金額損益": pnl_money,
                "口數": MTX_LOTS,
                "持有天數": 0,
            })
            position = None

    if not trades:
        return pd.DataFrame()

    trades_df = pd.DataFrame(trades)

    # 計算持有天數
    for i, row in trades_df.iterrows():
        d1 = pd.Timestamp(row["進場日"])
        d2 = pd.Timestamp(row["出場日"])
        trades_df.loc[i, "持有天數"] = (d2 - d1).days

    # 累計損益
    trades_df["累計損益"] = trades_df["金額損益"].cumsum()
    trades_df["累計點數"] = trades_df["點數損益"].cumsum()

    return trades_df


def compute_performance(trades_df: pd.DataFrame) -> dict:
    """計算績效統計"""
    if trades_df.empty:
        return {}

    total_trades = len(trades_df)
    wins = trades_df[trades_df["金額損益"] > 0]
    losses = trades_df[trades_df["金額損益"] <= 0]
    win_count = len(wins)
    loss_count = len(losses)
    win_rate = win_count / total_trades * 100

    total_pnl = trades_df["金額損益"].sum()
    avg_win = wins["金額損益"].mean() if len(wins) > 0 else 0
    avg_loss = losses["金額損益"].mean() if len(losses) > 0 else 0
    profit_factor = abs(wins["金額損益"].sum() / losses["金額損益"].sum()) if len(losses) > 0 and losses["金額損益"].sum() != 0 else float("inf")

    # 最大回撤
    cum = trades_df["累計損益"]
    peak = cum.cummax()
    drawdown = cum - peak
    max_dd = drawdown.min()

    # 月度績效
    trades_df["month"] = pd.to_datetime(trades_df["出場日"]).dt.to_period("M")
    monthly = trades_df.groupby("month").agg(
        trades=("金額損益", "count"),
        pnl=("金額損益", "sum"),
        wins=("金額損益", lambda x: (x > 0).sum()),
    ).reset_index()
    monthly["win_rate"] = (monthly["wins"] / monthly["trades"] * 100).round(1)

    return {
        "total_trades": total_trades,
        "win_count": win_count,
        "loss_count": loss_count,
        "win_rate": round(win_rate, 1),
        "total_pnl": total_pnl,
        "avg_win": round(avg_win, 0),
        "avg_loss": round(avg_loss, 0),
        "profit_factor": round(profit_factor, 2),
        "max_drawdown": max_dd,
        "monthly": monthly,
    }


# =============================================
#  STEP 5: Excel 輸出 (openpyxl 格式化)
# =============================================

def write_excel(df: pd.DataFrame, trades_df: pd.DataFrame, perf: dict):
    """產出格式化 Excel"""
    wb = Workbook()

    # 樣式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_fmt = "#,##0"
    pct_fmt = "0.0%"

    # ── Sheet 1: 籌碼紀錄 ──
    ws1 = wb.active
    ws1.title = "籌碼紀錄"
    cols1 = ["date", "futures_close", "近月", "遠月", "主力", "小散戶",
             "結算近-散戶", "籌碼", "背離", "信號"]
    headers1 = ["日期", "期貨收盤", "近月", "遠月", "主力", "小散戶",
                "結算近-散戶", "籌碼", "背離", "信號"]

    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    display_df = df[df["futures_close"].notna()].copy() if "futures_close" in df.columns else df.copy()
    for r_idx, (_, row) in enumerate(display_df.iterrows(), 2):
        for c_idx, col in enumerate(cols1, 1):
            val = row.get(col, "")
            if col == "date":
                val = str(val)[:10]
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if col in ["近月", "遠月", "主力", "小散戶", "結算近-散戶", "籌碼", "futures_close"]:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
            if col == "信號":
                if "翻多" in str(val):
                    cell.fill = green_fill
                elif "翻空" in str(val):
                    cell.fill = red_fill
                elif "💪" in str(val):
                    cell.fill = green_fill
                elif "⚠️" in str(val):
                    cell.fill = yellow_fill

    # 欄寬
    widths1 = [12, 10, 10, 10, 10, 10, 12, 10, 10, 18]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 小台回測 ──
    ws2 = wb.create_sheet("小台回測")
    headers2 = ["進場日", "進場價", "出場日", "出場價", "點數損益",
                "金額損益", "口數", "持有天數", "累計損益"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    if not trades_df.empty:
        for r_idx, (_, row) in enumerate(trades_df.iterrows(), 2):
            for c_idx, col in enumerate(headers2, 1):
                val = row.get(col, "")
                if col in ["進場日", "出場日"]:
                    val = str(val)[:10]
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if col in ["進場價", "出場價", "點數損益", "金額損益", "累計損益"]:
                    cell.number_format = num_fmt
                if col == "金額損益":
                    if val > 0:
                        cell.fill = green_fill
                    elif val < 0:
                        cell.fill = red_fill

    widths2 = [12, 10, 12, 10, 10, 12, 8, 10, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 月度績效 ──
    ws3 = wb.create_sheet("月度績效")
    headers3 = ["月份", "交易次數", "獲利金額", "勝率"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    if perf and "monthly" in perf:
        for r_idx, (_, row) in enumerate(perf["monthly"].iterrows(), 2):
            ws3.cell(row=r_idx, column=1, value=str(row["month"])).border = thin_border
            ws3.cell(row=r_idx, column=2, value=int(row["trades"])).border = thin_border
            c3 = ws3.cell(row=r_idx, column=3, value=row["pnl"])
            c3.number_format = num_fmt
            c3.border = thin_border
            if row["pnl"] > 0:
                c3.fill = green_fill
            elif row["pnl"] < 0:
                c3.fill = red_fill
            ws3.cell(row=r_idx, column=4, value=f"{row['win_rate']}%").border = thin_border

        # 總結行
        r_sum = len(perf["monthly"]) + 2
        ws3.cell(row=r_sum, column=1, value="總計").font = Font(bold=True)
        ws3.cell(row=r_sum, column=2, value=perf["total_trades"]).font = Font(bold=True)
        c_total = ws3.cell(row=r_sum, column=3, value=perf["total_pnl"])
        c_total.number_format = num_fmt
        c_total.font = Font(bold=True)
        ws3.cell(row=r_sum, column=4, value=f"{perf['win_rate']}%").font = Font(bold=True)

    for i, w in enumerate([12, 10, 14, 10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: 使用說明 ──
    ws4 = wb.create_sheet("使用說明")
    instructions = [
        ["小台指期貨籌碼策略 v2 — 使用說明"],
        [""],
        ["📊 資料來源"],
        ["  - 三大法人未平倉 (OI): FinMind API → TaiwanFuturesInstitutionalInvestors"],
        ["  - 大額交易人: 期交所 TAIFEX 網站爬取"],
        ["  - 台指期行情: FinMind API → TaiwanFuturesDaily"],
        ["  - 小台三大法人: FinMind API → MTX"],
        [""],
        ["📐 關鍵公式"],
        ["  - 遠月 = 三大法人全合約淨OI (inst_total，替代真正遠月合約OI)"],
        ["  - 小散戶 = -(小台法人淨未平倉) — 不除以4"],
        ["  - 主力 = 遠月/2 - 小散戶×2"],
        ["  - 籌碼 = 主力 - 主力(5天前) — 5日動量"],
        ["  - 信號: 翻多(主力由負翻正), 翻空(主力由正翻負)"],
        [""],
        ["🎯 交易規則"],
        ["  - 翻多信號日 → 隔天開盤買進小台1口"],
        ["  - 翻空信號日 → 隔天開盤賣出平倉"],
        ["  - 1口保證金: $93,500"],
        ["  - 來回成本: $114 (手續費+稅)"],
        ["  - 1點 = $50"],
        [""],
        ["⚠️ 注意事項"],
        ["  - 遠月用 inst_total 替代，與 Calvin 正確遠月有差異"],
        ["  - 待取得期交所近月/遠月分拆資料後可進一步修正"],
        ["  - 本策略使用「未平倉」(OI) 而非「買賣超」(交易量)"],
    ]
    for r, row in enumerate(instructions, 1):
        cell = ws4.cell(row=r, column=1, value=row[0] if row else "")
        if r == 1:
            cell.font = Font(bold=True, size=14)
        elif row and row[0].startswith("📊") or row and row[0].startswith("📐") or row and row[0].startswith("🎯") or row and row[0].startswith("⚠️"):
            cell.font = Font(bold=True, size=12)
    ws4.column_dimensions["A"].width = 65

    wb.save(str(EXCEL_FILE))
    print(f"\n  ✅ Excel 已儲存: {EXCEL_FILE}")


# =============================================
#  STEP 6: 終端輸出
# =============================================

def print_summary(df: pd.DataFrame, trades_df: pd.DataFrame, perf: dict):
    """印出終端摘要"""

    # === 公式驗證 (5天) ===
    print("\n" + "=" * 70)
    print("  公式驗證：主力 = 遠月/2 - 小散戶×2")
    print("  (遠月 = inst_total 替代版)")
    print("=" * 70)

    validation = [
        ("2023-12-27", -6084, -13002, 22962),
        ("2023-12-28", -3428, -7039, 12364),
        ("2023-12-29", -9820, -4242, 3574),
        ("2024-01-02", -6718, 3743, -10845),
        ("2024-01-03", -30190, 18654, -52403),
    ]

    print(f"  {'日期':<12} {'遠月':>8} {'小散戶':>8} {'算出主力':>10} {'正確答案':>10} {'結果':>6}")
    print(f"  {'─'*58}")
    all_pass = True
    for date_str, correct_yuan, correct_retail, correct_zhuli in validation:
        row = df[df["date"].astype(str).str[:10] == date_str]
        if row.empty:
            print(f"  {date_str:<12} — 資料不存在")
            all_pass = False
            continue
        r = row.iloc[0]
        my_yuan = int(r.get("遠月", 0))
        my_retail = int(r.get("小散戶", 0))
        my_zhuli = int(r.get("主力", 0))

        yuan_match = "✅" if my_yuan == correct_yuan else f"≠{correct_yuan:+,}"
        retail_match = "✅" if my_retail == correct_retail else f"≠{correct_retail:+,}"
        zhuli_match = "✅" if my_zhuli == correct_zhuli else "❌"
        if my_zhuli != correct_zhuli:
            all_pass = False

        print(f"  {date_str:<12} {my_yuan:>+8,} {my_retail:>+8,} {my_zhuli:>+10,} {correct_zhuli:>+10,} {zhuli_match}")
        if my_zhuli != correct_zhuli:
            print(f"    ↳ 遠月差異: 本系統={my_yuan:+,} vs 正確={correct_yuan:+,} (差{my_yuan-correct_yuan:+,})")
            print(f"    ↳ 小散戶差異: 本系統={my_retail:+,} vs 正確={correct_retail:+,} (差{my_retail-correct_retail:+,})")

    if not all_pass:
        print(f"\n  ⚠ 主力值有差異（因遠月用 inst_total 替代，非真正遠月合約OI）")
        print(f"  ⚠ 小散戶 = -mtx_inst_net_oi 已驗證正確")
        print(f"  ⚠ 待取得期交所近月/遠月分拆資料後可修正")

    # === 最新資料 ===
    if not df.empty and "futures_close" in df.columns:
        latest = df.dropna(subset=["futures_close"]).iloc[-1]
        print(f"\n{'─'*70}")
        print(f"  最新日期: {str(latest['date'])[:10]}")
        print(f"  期貨收盤: {latest['futures_close']:,.0f}")
        print(f"  遠月(inst_total): {int(latest.get('遠月', 0)):+,}")
        print(f"  小散戶(-mtx_inst): {int(latest.get('小散戶', 0)):+,}")
        print(f"  主力(遠月/2-散戶×2): {int(latest.get('主力', 0)):+,}")
        print(f"  籌碼(5日動量): {int(latest.get('籌碼', 0)):+,}")
        print(f"  信號: {latest.get('信號', '')}")

    # === 回測績效 ===
    if perf:
        print(f"\n{'─'*70}")
        print("  📈 小台回測績效")
        print(f"{'─'*70}")
        print(f"  交易次數: {perf['total_trades']} 筆")
        print(f"  勝率: {perf['win_rate']}% ({perf['win_count']}勝 / {perf['loss_count']}敗)")
        print(f"  累計點數: {int(trades_df['點數損益'].sum()) if not trades_df.empty else 0:+,}")
        print(f"  累計金額: ${perf['total_pnl']:+,.0f}")
        print(f"  平均獲利: ${perf['avg_win']:+,.0f}")
        print(f"  平均虧損: ${perf['avg_loss']:+,.0f}")
        print(f"  獲利因子: {perf['profit_factor']:.2f}")
        print(f"  最大回撤: ${perf['max_drawdown']:+,.0f}")

    print("\n" + "=" * 70)


# =============================================
#  主程式
# =============================================

def fetch_all_data(start_date: str, end_date: str) -> pd.DataFrame:
    """抓取所有資料並整合"""
    print(f"\n{'='*60}")
    print(f"  資料抓取: {start_date} ~ {end_date}")
    print(f"{'='*60}\n")

    # 1. 三大法人台指期 OI
    df_inst_raw = fetch_futures_institutional(start_date, end_date)
    time.sleep(REQUEST_DELAY)

    # 2. 台指期行情
    df_daily = fetch_futures_daily(start_date, end_date)
    time.sleep(REQUEST_DELAY)

    # 3. 小台三大法人 OI
    df_mini_raw = fetch_mini_futures_institutional(start_date, end_date)
    time.sleep(REQUEST_DELAY)

    # 整理三大法人
    df_inst = process_institutional(df_inst_raw)
    df_mini = process_mini_institutional(df_mini_raw)

    # 合併
    if df_inst.empty:
        print("  ❌ 三大法人資料為空，無法繼續")
        return pd.DataFrame()

    df = df_inst.copy()
    if not df_daily.empty:
        df = df.merge(df_daily, on="date", how="left")
    if not df_mini.empty:
        df = df.merge(df_mini, on="date", how="left")

    # 4. 大額交易人 (TAIFEX爬蟲)
    all_dates = sorted(df["date"].unique())
    print(f"\n  共 {len(all_dates)} 個交易日，開始爬大額交易人...")
    df_lt = batch_large_traders(all_dates)
    if not df_lt.empty:
        df = df.merge(df_lt, on="date", how="left")
        print(f"  ✅ 大額交易人: {len(df_lt)} 天")
    else:
        print("  ⚠ 大額交易人資料為空")

    # 儲存 parquet
    df.to_parquet(str(PARQUET_FILE), index=False)
    print(f"\n  ✅ 原始資料已儲存: {PARQUET_FILE} ({len(df)} 天)")

    return df


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--verify", action="store_true", help="只跑驗證")
    parser.add_argument("--update", action="store_true", help="每日更新")
    parser.add_argument("--start", default="2023-01-01", help="起始日期")
    parser.add_argument("--end", default="", help="結束日期 (預設今天)")
    args = parser.parse_args()

    if not args.end:
        args.end = datetime.now().strftime("%Y-%m-%d")

    print(f"\n{'#'*60}")
    print(f"  小台指期貨籌碼策略 v1.0")
    print(f"  {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"{'#'*60}")

    if args.verify:
        # 只驗證 2024-01-29 附近的資料
        print("\n  🔍 驗證模式: 抓取 2024-01-25 ~ 2024-02-02")
        df = fetch_all_data("2024-01-25", "2024-02-02")
        if not df.empty:
            df = compute_derived(df)
            print_summary(df, pd.DataFrame(), {})
        return

    if args.update and PARQUET_FILE.exists():
        # 每日更新模式
        df_existing = pd.read_parquet(PARQUET_FILE)
        last_date = df_existing["date"].max()
        print(f"\n  📅 更新模式: 從 {last_date} 到 {args.end}")
        df_new = fetch_all_data(str(last_date), args.end)
        if not df_new.empty:
            df = pd.concat([df_existing, df_new], ignore_index=True)
            df = df.drop_duplicates(subset=["date"], keep="last")
            df = df.sort_values("date").reset_index(drop=True)
        else:
            df = df_existing
    else:
        # 完整抓取
        df = fetch_all_data(args.start, args.end)

    if df.empty:
        print("  ❌ 無資料")
        return

    # 計算衍生欄位
    df = compute_derived(df)

    # 回測
    trades_df = run_backtest(df)
    perf = compute_performance(trades_df) if not trades_df.empty else {}

    # Excel
    write_excel(df, trades_df, perf)

    # 終端摘要
    print_summary(df, trades_df, perf)

    # 儲存完整資料
    df.to_parquet(str(PARQUET_FILE), index=False)

    print(f"\n  📁 檔案位置:")
    print(f"     Excel: {EXCEL_FILE}")
    print(f"     Data:  {PARQUET_FILE}")
    print(f"\n  完成！")


if __name__ == "__main__":
    main()
